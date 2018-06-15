import openpyxl
import csv

wb = openpyxl.load_workbook('../../../ragd/PowerPlantSourceData/GlobalData_Solar_CPV_180427.xlsx')
# Get the active sheet
ws = wb.active  

# Retrieving the country-region lookup values as a key-value pair
wb1 = openpyxl.load_workbook('../../../ragd/PowerPlantDataFields-Source Matrix_Final.xlsx')
ws1 = wb1["Country-Region Lookup"]
# Retrieving the source-variables titles as a list of data.
ws2 = wb1["Source - Variables Matrix"]

# Headers from the excel file.
excel_header = []

# We used Source Matrix File to populate countries names as the keys and regions as the values into the dictionary.
countries = {}
for row_number,row in enumerate(ws1.rows):
    if row_number != 0:
        countries[row[1].value]=row[0].value

# This dic is to be able to write the months as integer values in the csv file.
months = {'Jan':1, 'Feb':2, 'Mar':3, 'Apr':4, 'May':5, 'Jun':6, 'Jul':7, 'Aug':8, 'Sep':9, 'Oct':10, 'Nov':11, 'Dec':12}

# this def will write most of the rows
def clean_data_values(row, row_number, column_number):
    return ws.cell(row=row_number + 1, column=column_number + 1).value

# This will write plant capacity row and it's connected with get_plant_capacity_value
def clean_plant_capacity_value(row, row_number, column_number):
    capacity_value = get_plant_capacity_value(row, excel_header)
    return capacity_value

# This will write the project capatiy and it's connected to get_project_capacity value
def clean_project_capacity_value(row, row_number, column_number):
    project_value = get_project_capacity_value(row, excel_header)
    return project_value

def clean_month_value(row, row_number, column_number):
    month_int_value = get_month_value(row, excel_header)
    return month_int_value

def clean_region_value(row, row_number, column_number):
    regions_value = get_regions_values(row, excel_header)
    return regions_value

# The ones that are returning just a string without any other logic are values that were added by default.
def clean_infr_name(row, row_number, column_number):
    return "Power Plant"

def clean_fuel_name(row, row_number, column_number):
    return "Solar"

def clean_pfuel_name(row, row_number, column_number):
    return "Solar CPV"

def plant_capacity_unit(row, row_number, column_number):
    return "MW"

def co2_unit(row, row_number, column_number):
    return "Tons per annum"

def currency_code(row, row_number, column_number):
    return "USD"

def bool_case(row, row_number, column_number):
    return "TRUE"

def output_unit(row, row_number, column_number):
    return "MWh"

# This is for values that needs to be empty or N/A in the matrix sheet
def clean_empty_name(row, row_number, column_number):
    return ""

def get_plant_capacity_value(row, excel_header):
    plant_capacity_number = excel_header.index('Plant Capacity (MW)')
    total_capacity_number = excel_header.index('Total Capacity (MW)')
    if row[plant_capacity_number].value:
        plant_value = row[plant_capacity_number].value
    else:
        plant_value = row[total_capacity_number].value
    
    return plant_value

def get_project_capacity_value(row, excel_header):
    active_number = excel_header.index('Active Capacity (MW)')
    pipeline_number = excel_header.index('Pipeline Capacity (MW)')
    discontinued_number = excel_header.index('Discontinued Capacity (MW)')

    if row[active_number].value:
        project_value = row[active_number].value
    elif row[pipeline_number].value:
        project_value = row[pipeline_number].value
    else:
        project_value = row[discontinued_number].value
    
    return project_value

def get_month_value(row, excel_header):
    month_number = excel_header.index('Month Online')
    # import ipdb; ipdb.set_trace()
    if row[month_number].value in months.keys() and not row[month_number].font.italic:
        month_number = months.get(row[month_number].value)
        return month_number
    else:
        return ""

def get_regions_values(row, excel_header):
    country_number = excel_header.index('Country')
    region_value = countries.get(row[country_number].value)
    return region_value

# Function that let us skip the entire row if some conditions are true.
def skip_row(row, excel_header):
    bool_value = False
    country_number = excel_header.index('Country')
    deco_year_number = excel_header.index('Decommissioning Year')

    if row[country_number].value not in countries.keys():
        bool_value = True

    if row[deco_year_number].value and row[deco_year_number].value <= 2006: 
        bool_value = True
    
    plant_value = get_plant_capacity_value(row, excel_header)

    if plant_value and plant_value < 100:
        bool_value = True

    return bool_value

headers = (
    ("Power Plant Name","Power Plant Name",clean_data_values),
    ("Infrastructure Type","",clean_infr_name),
    ("Country","Country",clean_data_values),
    ("Region","", clean_region_value), # This should be those that are in the countries dic.
    ("Plant Status","Status",clean_data_values), # This will need something to set the final status.
    ("Plant Day Online","", clean_empty_name), # Remove italiziced values
    ("Plant Month Online","",clean_month_value),
    ("Plant Year Online","Year Online", clean_data_values), # One final output - Remove italicized values
    ("Decommissioning Day","", clean_empty_name),
    ("Decommissioning Month","", clean_empty_name),
    ("Decommissioning Year","Decommissioning Year", clean_data_values), # One final output
    ("Owner 1","Owner", clean_data_values), # Standardize names based on the organization tab if applicable 
    ("Owner 1 Stake","Owner Stake", clean_data_values),
    ("Owner 2","", clean_empty_name), # There will be multiple owners values just in one row.
    ("Owner 2 Stake","", clean_empty_name),
    ("Plant Fuel 1","", clean_fuel_name),
    ("Plant Fuel 2","", clean_empty_name),
    ("Plant Fuel 3","", clean_empty_name),
    ("Plant Fuel 4","", clean_empty_name),
    ("Project Fuel 1","", clean_pfuel_name),
    ("Project Fuel 2","", clean_empty_name),
    ("Project Fuel 3","", clean_empty_name),
    ("Project Fuel 4","", clean_empty_name),
    ("Plant Capacity","", clean_plant_capacity_value),
    ("Plant Capacity Unit","", plant_capacity_unit),
    ("Project Capacity","", clean_project_capacity_value),
    ("Project Capacity Unit","", plant_capacity_unit),
    ("Plant Output","", clean_empty_name), # This has to be multiplied * 1000 for the final output 
    ("Plant Output Unit","", output_unit),
    ("Plant Output Year","", clean_empty_name),
    ("Estimated Plant Output","Average Output", clean_data_values), # Just one output and just the int value * 1000, Remove if plant output is NOT null
    ("Estimated Plant Output Unit","", output_unit), # this will be always MWh
    ("Project Output","", clean_empty_name), # Just one output and just int * 1000
    ("Project Output Unit","", output_unit),
    ("Project Output Year","", clean_empty_name), # Only keep the most recent avaialble year's data
    ("Estimated Project Output","Average Output", clean_data_values), # Just one output and just the int value * 1000, Remove if plant output is NOT null
    ("Estimated Project Output Unit","", output_unit),
    ("Plant CO2 Emissions","CO2 Emissions (Tonnes per annum)", clean_data_values), # This will be just the int value 
    ("Plant CO2 Emissions Unit","", co2_unit),
    ("Project CO2 Emissions","CO2 Emissions (Tonnes per annum)", clean_data_values), # This will be just the int value 
    ("Project CO2 Emissions Unit","", co2_unit),
    ("Grid Connected","", clean_empty_name),
    ("Manufacturer 1","", clean_empty_name), # Standardize names based on the organization tab if applicable 
    ("Manufacturer 2","", clean_empty_name),
    ("Manufacturer 3","", clean_empty_name),
    ("Manufacturer 4","", clean_empty_name),
    ("Manufacturer 5","", clean_empty_name),
    ("SOx Reduction System","", clean_empty_name),
    ("NOx Reduction System","", clean_empty_name),
    ("Project Name","Subsidiary Asset Name", clean_data_values), # Consolidate & Remove Duplicates
    ("Project Status","Status", clean_data_values), # See Status Conversions Chart
    ("Total Cost","Capex USD", clean_data_values), # Remove italicized values - Convert to integer - move text
    ("Total Cost Currency","", currency_code),
    ("Start Day","", clean_empty_name),
    ("Start Month","", clean_empty_name),
    ("Start Year","", clean_empty_name),
    ("Construction Start Day","", clean_empty_name),
    ("Construction Start Month","", clean_empty_name),
    ("Construction Start Year","", clean_empty_name),
    ("Completion Day","", clean_empty_name),
    ("Completion Month","", clean_month_value),
    ("Completion Year","Year Online", clean_data_values), # Remove row < 2006
    ("New Construction","", bool_case),
    ("Latitude","Latitude", clean_data_values),
    ("Longitude","Longitude", clean_data_values),
    ("Initiative","", clean_empty_name),
    ("Contractor 1","EPC Contractor", clean_data_values), # Move values into separate columns 
    ("Contractor 2","", clean_empty_name),
    ("Contractor 3","", clean_empty_name),
    ("Contractor 4","", clean_empty_name),
    ("Contractor 5","", clean_empty_name),
    ("Contractor 6","", clean_empty_name),
    ("Contractor 7","", clean_empty_name),
    ("Contractor 8","", clean_empty_name),
    ("Contractor 9","", clean_empty_name),
    ("Contractor 10","", clean_empty_name),
    ("Contractor 11","", clean_empty_name),
    ("Contractor 12","", clean_empty_name),
    ("Consultant","", clean_empty_name),
    ("Implementing Agency","", clean_empty_name),
    ("Operator 1","Operator", clean_data_values), # Standardize names based on the organization tab if applicable 
    ("Operator 2","", clean_empty_name),
    ("Funder 1","",clean_empty_name),
    ("Funding Amount 1","", clean_empty_name),
    ("Funding Currency 1","", clean_empty_name),
    ("Funder 2","", clean_empty_name),
    ("Funding Amount 2","", clean_empty_name),
    ("Funding Currency 2","", clean_empty_name),
    ("Sources","", clean_empty_name),
    ("Notes","Asset Notes", clean_data_values)
)

# A list with the output headers retrieved from the header tuple.
titles = [header[0] for header in headers]

# Populate the csv file with the excel data.
with open('test_tuple.csv', 'w') as csv_file:
    writer = csv.DictWriter(csv_file, fieldnames=titles)
    writer.writeheader()

    # Iterating trough the entire row getting the row number and row value
    for row_number,row in enumerate(ws.rows):
        row_data = {}
        if row_number == 0:
            excel_header = excel_header + [cell.value for cell in row]

        if row_number > 0 and not skip_row(row, excel_header):
            for title, excel_title, clean_up_function in headers:
                column_number = 0
                if excel_title:
                    column_number = excel_header.index(excel_title)
                row_data[title] = clean_up_function(row, row_number, column_number)
            writer.writerow(row_data)
    