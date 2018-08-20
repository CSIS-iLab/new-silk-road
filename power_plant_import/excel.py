
import openpyxl, re, csv, cchardet, io
from collections import OrderedDict


def load_workbook_data(filepath, headers=True, break_on_blank=True):
    """get workbook data from the given file. 
    Arguments:

    * filepath: the Excel filesystem path
    * headers=True: indicates the presence of a header row.
        - if True, use the values in the first row as record keys.
        - if False, use the Excel names of the columns as record keys.
    * break_on_blank=True: if True, stop processing at the first empty row.

    Returns an OrderedDict containing the worksheets (in order), 
    with each worksheet as a list of OrderedDict records, each row as a record.
    (We're using OrderedDict so that worksheets and columns can be accessed by index if desired.)
    """

    def make_record(keys, row):
        record = OrderedDict()
        for i in range(len(row)):
            value = row[i].value
            if type(value)==str:
                value = re.sub(r"\s+", " ", value).strip()
            record[keys[i]] = value
        return record

    workbook = openpyxl.load_workbook(filepath)
    wbdata = OrderedDict()
    for sheet in workbook._sheets:
        title = sheet.title.strip()
        wbdata[title] = []
        rows = sheet.rows.__iter__()  # sheet.rows doesn't behave as a normal generator, so force it
        row = rows.__next__()  # <= this consumes the row
        if headers == True:
            keys = [cell.value for cell in row]
        else:
            keys = [excel_key(i) for i in range(len(row))]
            record = make_record(keys, row)
            wbdata[title].append(record)
        keys = [re.sub(r"\s+", " ", key or '').strip() for key in keys]
        for row in rows:
            record = make_record(keys, row)
            if break_on_blank == True and set(record.values()) == {None}:
                break
            wbdata[title].append(record)
    return wbdata


def load_csv(filename, dialect="excel", encoding=None, headings=True):
    """load data from an Excel CSV file (comma delimiter)"""
    with open(filename, "rb") as f:
        fd = f.read()
    if encoding is None:
        encoding = cchardet.detect(fd)["encoding"]
    txt = fd.decode(encoding)
    reader = csv.reader(io.StringIO(txt), dialect=dialect)
    csvdata = []
    if headings == True:
        # the first row is the keys
        keys = reader.__next__()
    else:
        # the first row is data, keys are letters
        row = reader.__next__()
        keys = [excel_key(i) for i in range(len(row))]
        d = OrderedDict(
            **{keys[i].strip(): re.sub(r"\s+", " ", row[i].strip()) for i in range(len(row))}
        )
        csvdata.append(d)
    for row in reader:
        d = OrderedDict(
            **{keys[i].strip(): re.sub(r"\s+", " ", row[i].strip()) for i in range(len(row))}
        )
        csvdata.append(d)
    return csvdata


def excel_key(index):
    """create a key for index by converting index into a base-26 number, using A-Z as the characters."""
    X = lambda n: ~n and X((n // 26) - 1) + chr(65 + (n % 26)) or ""  # chr(65)=='A'
    return X(int(index))


def worksheet_dict(worksheet_data, primary_key, merge=True, report=True):
    """convert a worksheet_data list into a dict, 
    using the primary_key (str key or tuple of record keys) to create dict keys.
    Returns an OrderedDict.
    raises a ValueError if a duplicate primary key is found
    """
    wsdict = OrderedDict()
    for record in worksheet_data:
        # key can either be tuple or string, normalize it (no whitespace)
        if isinstance(primary_key, str):
            key = record.get(primary_key).strip()
        else:
            key = tuple([str(record.get(key)).strip() for key in primary_key])
        if key not in wsdict:
            wsdict[key] = record
        elif merge != True and report==True:
            print(f"Duplicate key: {key}")
        else:
            if report==True:
                print(f"Duplicate for {primary_key} = {key}:")
            for field, val in record.items():
                if val is not None:
                    if wsdict[key][field] is None:
                        wsdict[key][field] = val
                    elif wsdict[key][field] != val and report==True:
                        print(f'  CONFLICT: {field} = {wsdict[key][field]}')
                        print(f'            {field} = {val}')
    return wsdict


def value_to_float(value):
    """convert a value as given to a float"""
    md = re.search(r"([\d\.,]+)", value or "", flags=re.I)
    if md is None:
        return
    else:
        number = float(md.group(0).replace(",", ""))
        if "million" in value.lower():
            number *= 1e6
        elif "billion" in value.lower():
            number *= 1e9
        return number


if __name__ == "__main__":
    import os, sys, json

    with_italics = {}
    filenames = sys.argv[1:]
    print(len(filenames), "files")
    for filename in filenames:
        print(filenames.index(filename) + 1, filename)
        ext = os.path.splitext(filename)[-1].lower()
        try:
            if ext=='.xlsx':
                wbdata = load_workbook_data(filename)
                worksheets = {title:wbdata[title] for title in wbdata.keys()}
            elif ext=='.csv':
                title = os.path.splitext(os.path.basename(filename))[0]
                worksheet_data = load_csv(filename)
                worksheets = {title:worksheet_data}
            for title in worksheets.keys():
                worksheet_data = worksheets[title]
                with open(f"{os.path.splitext(filename)[0]}_{title}.json", "w") as f:
                    f.write(json.dumps(worksheet_data, indent=2))
        except:
            print(os.path.basename(filename), ":", sys.exc_info()[1])
